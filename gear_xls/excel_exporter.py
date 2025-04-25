#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Модуль для создания Excel-файла на основе данных из HTML-версии расписания.
Принимает данные о блоках занятий и формирует Excel-файл в том же формате,
в котором исходный файл был загружен для создания HTML-расписания.
"""

import os
import json
import logging
import tempfile
from datetime import datetime, time
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger('excel_exporter')

def create_excel_from_html_data(schedule_data, output_file=None):
    """
    Создает Excel-файл на основе данных расписания из HTML-версии.
    
    Args:
        schedule_data (list): Список словарей с данными о занятиях
        output_file (str, optional): Путь к выходному файлу. 
                                    Если None, создает имя по шаблону "schedule_export_YYYY-MM-DD.xlsx"
    
    Returns:
        str: Путь к созданному Excel-файлу
    """
    if not schedule_data:
        logger.error("Нет данных для экспорта в Excel")
        return None
    
    # Если имя выходного файла не указано, создаем имя по умолчанию
    if not output_file:
        current_date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        output_file = f"schedule_export_{current_date}.xlsx"
    
    # Создаем каталог для выходного файла, если его нет
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    logger.info(f"Начинаем создание Excel-файла: {output_file}")
    logger.info(f"Количество записей для экспорта: {len(schedule_data)}")
    
    try:
        # Создаем новую книгу Excel и активный лист
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"
        
        # Добавляем заголовки
        headers = ["Занятие", "Группа", "Преподаватель", "Кабинет", 
                  "Здание", "День", "Начало", "Конец", "Продолжительность"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Заполняем данными
        for row_idx, activity in enumerate(schedule_data, 2):
            # Предмет
            ws.cell(row=row_idx, column=1).value = activity.get('subject', '')
            # Группа
            ws.cell(row=row_idx, column=2).value = activity.get('students', '')
            # Преподаватель
            ws.cell(row=row_idx, column=3).value = activity.get('teacher', '')
            # Кабинет
            ws.cell(row=row_idx, column=4).value = activity.get('room', '')
            # Здание
            ws.cell(row=row_idx, column=5).value = activity.get('building', '')
            # День
            ws.cell(row=row_idx, column=6).value = activity.get('day', '')
            
            # Время начала (преобразуем в строку формата HH:MM)
            start_time = activity.get('start_time', '00:00')
            try:
                # Если start_time является словарем
                if isinstance(start_time, dict):
                    # Проверяем наличие ключей hour и minute
                    if 'hour' in start_time and 'minute' in start_time:
                        hour = int(start_time.get('hour', 0))
                        minute = int(start_time.get('minute', 0))
                        formatted_start_time = f"{hour:02d}:{minute:02d}"
                    else:
                        # Для других форматов словаря, преобразуем в строку
                        formatted_start_time = f"{start_time}"
                # Если start_time уже строка с правильным форматом (HH:MM)
                elif isinstance(start_time, str) and ':' in start_time:
                    # Оставляем как есть, это уже правильный формат
                    formatted_start_time = start_time
                # Преобразуем time object в строку
                elif hasattr(start_time, 'hour') and hasattr(start_time, 'minute'):
                    formatted_start_time = f"{start_time.hour:02d}:{start_time.minute:02d}"
                # Для всех других случаев, преобразуем в строку
                else:
                    formatted_start_time = str(start_time)
                
                # Записываем отформатированное время в ячейку
                ws.cell(row=row_idx, column=7).value = formatted_start_time
            except Exception as e:
                logger.warning(f"Ошибка при обработке времени начала: {e}, строка {row_idx}")
                ws.cell(row=row_idx, column=7).value = str(start_time)
            
            # Время окончания (преобразуем в строку формата HH:MM)
            end_time = activity.get('end_time', '00:00')
            try:
                # Если end_time является словарем
                if isinstance(end_time, dict):
                    # Проверяем наличие ключей hour и minute
                    if 'hour' in end_time and 'minute' in end_time:
                        hour = int(end_time.get('hour', 0))
                        minute = int(end_time.get('minute', 0))
                        formatted_end_time = f"{hour:02d}:{minute:02d}"
                    else:
                        # Для других форматов словаря, преобразуем в строку
                        formatted_end_time = f"{end_time}"
                # Если end_time уже строка с правильным форматом (HH:MM)
                elif isinstance(end_time, str) and ':' in end_time:
                    # Оставляем как есть, это уже правильный формат
                    formatted_end_time = end_time
                # Преобразуем time object в строку
                elif hasattr(end_time, 'hour') and hasattr(end_time, 'minute'):
                    formatted_end_time = f"{end_time.hour:02d}:{end_time.minute:02d}"
                # Для всех других случаев, преобразуем в строку
                else:
                    formatted_end_time = str(end_time)
                
                # Записываем отформатированное время в ячейку
                ws.cell(row=row_idx, column=8).value = formatted_end_time
            except Exception as e:
                logger.warning(f"Ошибка при обработке времени окончания: {e}, строка {row_idx}")
                ws.cell(row=row_idx, column=8).value = str(end_time)
            
            # Продолжительность в минутах
            ws.cell(row=row_idx, column=9).value = activity.get('duration', 0)
            
            # Применяем стили к ячейкам
            for col_idx in range(1, 10):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Авто-ширина для столбцов
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Сохраняем файл
        wb.save(output_file)
        logger.info(f"Excel-файл успешно создан: {output_file}")
        
        return output_file
    
    except Exception as e:
        logger.error(f"Ошибка при создании Excel-файла: {e}")
        import traceback
        traceback.print_exc()
        return None

def process_schedule_export_request(request_data, output_dir="excel_exports"):
    """
    Обрабатывает запрос на экспорт расписания в Excel.
    
    Args:
        request_data (str): JSON-строка с данными расписания
        output_dir (str, optional): Директория для сохранения выходного файла
        
    Returns:
        str: Путь к созданному Excel-файлу
    """
    try:
        # Создаем директорию для экспорта, если её нет
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # Преобразуем JSON-строку в Python-объект
        if isinstance(request_data, str):
            try:
                schedule_data = json.loads(request_data)
                logger.info(f"JSON успешно распарсен, получен объект типа: {type(schedule_data)}")
            except json.JSONDecodeError as json_err:
                logger.error(f"Ошибка декодирования JSON: {json_err}")
                logger.debug(f"Полученные данные: {request_data[:100]}...")
                return None
        else:
            schedule_data = request_data
            logger.info(f"Получены данные в формате {type(schedule_data)}")
        
        # Проверяем валидность данных
        if not isinstance(schedule_data, list):
            logger.error(f"Неверный формат данных. Ожидается список, получено: {type(schedule_data)}")
            return None
        
        if not schedule_data:
            logger.error("Список занятий пуст")
            return None
        
        # Проверяем наличие необходимых полей в первой записи
        required_fields = ['subject', 'day', 'start_time', 'end_time']
        first_record = schedule_data[0]
        
        # Выводим информацию о первой записи для отладки
        logger.info(f"Первая запись: {first_record}")
        
        missing_fields = [field for field in required_fields if field not in first_record]
        
        if missing_fields:
            logger.error(f"В данных отсутствуют обязательные поля: {', '.join(missing_fields)}")
            return None
        
        # Проверяем форматы полей времени в первой записи
        start_time = first_record.get('start_time')
        end_time = first_record.get('end_time')
        
        logger.info(f"Тип start_time: {type(start_time)}, значение: {start_time}")
        logger.info(f"Тип end_time: {type(end_time)}, значение: {end_time}")
        
        logger.info(f"Валидация данных успешно пройдена. Получено {len(schedule_data)} записей")
        
        # Формируем имя выходного файла
        current_date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        output_file = os.path.join(output_dir, f"schedule_export_{current_date}.xlsx")
        
        # Создаем Excel-файл
        return create_excel_from_html_data(schedule_data, output_file)
    
    except Exception as e:
        logger.error(f"Ошибка при обработке запроса на экспорт: {e}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    # Пример использования
    sample_data = [
        {
            "subject": "Математика",
            "students": "2a",
            "teacher": "Иванов И.И.",
            "room": ".12",
            "room_display": ".12",
            "building": "Kolibri",
            "day": "Mo",
            "start_time": "09:00",
            "end_time": "09:45",
            "duration": 45,
            "color": "#FFD700"
        }
    ]
    
    test_file = "test_export.xlsx"
    create_excel_from_html_data(sample_data, test_file)