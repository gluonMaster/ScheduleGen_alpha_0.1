#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Модуль для парсинга Excel-файла с расписанием.
Извлекает информацию о занятиях, преподавателях, кабинетах и группах.
"""

import json
import logging
import pandas as pd
import openpyxl
from datetime import datetime
from typing import Dict, Any

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger('excel_parser')


def _build_header_lookup(sheet):
    lookup = {}
    for col_idx in range(1, sheet.max_column + 1):
        value = sheet.cell(row=1, column=col_idx).value
        if value is None:
            continue
        lookup[str(value).strip().lower()] = col_idx
    return lookup


def _resolve_column_index(header_lookup, default_index, *header_names):
    for header_name in header_names:
        idx = header_lookup.get(str(header_name).strip().lower())
        if idx is not None:
            return idx
    return default_index

def parse_schedule(excel_file):
    """
    Парсит Excel-файл с расписанием и возвращает словарь,
    где ключ – уникальный идентификатор активности, а значение – словарь с информацией:
      - day: день активности
      - start_time: время начала (в формате 'HH:MM')
      - end_time: время окончания (в формате 'HH:MM')
      - duration: продолжительность в минутах
      - teacher: имя преподавателя
      - subject: название активности
      - room: название кабинета
      - room_display: отображаемое название кабинета (без префикса здания)
      - building: название здания
      - students: название группы
      
    Args:
        excel_file (str): Путь к Excel-файлу с расписанием
        
    Returns:
        dict: Словарь с информацией о занятиях
    """
    logger.info(f"Начинаем парсинг Excel-файла: {excel_file}")
    activities = {}
    
    try:
        # Загружаем Excel-файл с помощью openpyxl (для корректной работы с датами и временем)
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = workbook['Schedule']
        
        header_lookup = _build_header_lookup(sheet)
        subject_col = _resolve_column_index(header_lookup, 1, "subject")
        group_col = _resolve_column_index(header_lookup, 2, "group")
        teacher_col = _resolve_column_index(header_lookup, 3, "teacher")
        room_col = _resolve_column_index(header_lookup, 4, "room")
        building_col = _resolve_column_index(header_lookup, 5, "building")
        day_col = _resolve_column_index(header_lookup, 6, "day")
        start_time_col = _resolve_column_index(header_lookup, 7, "start_time")
        end_time_col = _resolve_column_index(header_lookup, 8, "end_time")
        duration_col = _resolve_column_index(header_lookup, 9, "duration")
        lesson_type_col = _resolve_column_index(
            header_lookup,
            10 if sheet.max_column >= 10 else None,
            "lesson_type", "тип занятия",
        )
        trial_dates_col = _resolve_column_index(
            header_lookup,
            11 if sheet.max_column >= 11 else None,
            "trial_dates_json", "даты (json)",
        )

        # Начинаем со второй строки (первая - заголовки)
        row_idx = 2
        act_id = 1
        
        while sheet.cell(row=row_idx, column=subject_col).value is not None:
            subject = sheet.cell(row=row_idx, column=subject_col).value
            group = sheet.cell(row=row_idx, column=group_col).value
            teacher = sheet.cell(row=row_idx, column=teacher_col).value
            room = sheet.cell(row=row_idx, column=room_col).value
            building = sheet.cell(row=row_idx, column=building_col).value
            day = sheet.cell(row=row_idx, column=day_col).value
            
            # Извлекаем время начала и окончания с корректной обработкой
            start_time_cell = sheet.cell(row=row_idx, column=start_time_col)
            end_time_cell = sheet.cell(row=row_idx, column=end_time_col)
            start_time = extract_time_from_cell(start_time_cell)
            end_time = extract_time_from_cell(end_time_cell)
            
            # Извлекаем продолжительность
            duration = sheet.cell(row=row_idx, column=duration_col).value
            lesson_type = None
            trial_dates = []

            if lesson_type_col is not None:
                lesson_type_value = sheet.cell(row=row_idx, column=lesson_type_col).value
                if lesson_type_value is not None:
                    lesson_type = str(lesson_type_value).strip().lower() or None

            if lesson_type == "trial" and trial_dates_col is not None:
                raw_trial_dates = sheet.cell(row=row_idx, column=trial_dates_col).value
                if raw_trial_dates not in (None, ""):
                    try:
                        parsed_trial_dates = json.loads(raw_trial_dates)
                        if isinstance(parsed_trial_dates, list):
                            trial_dates = [str(item) for item in parsed_trial_dates if item is not None]
                        else:
                            logger.warning(
                                "trial_dates_json is not a list in row %s: %r",
                                row_idx,
                                raw_trial_dates,
                            )
                    except Exception as exc:
                        logger.warning(
                            "Failed to parse trial_dates_json in row %s: %s",
                            row_idx,
                            exc,
                        )
            
            # Обработка названия кабинета и извлечение здания
            room_info = process_room_name(room, building)
            
            # Добавляем активность в словарь
            activities[act_id] = {
                "day": day,
                "start_time": start_time,
                "end_time": end_time,
                "duration": int(duration) if duration is not None else 0,
                "teacher": teacher,
                "subject": subject,
                "room": room_info["full_name"],
                "room_display": room_info["display_name"],
                "building": room_info["building"],
                "students": group,
            }
            if lesson_type:
                activities[act_id]["lesson_type"] = lesson_type
            if lesson_type == "trial":
                activities[act_id]["trial_dates"] = trial_dates
            
            act_id += 1
            row_idx += 1
            
        logger.info(f"Парсинг завершен. Извлечено {len(activities)} занятий.")
        return activities
        
    except Exception as e:
        logger.error(f"Ошибка при парсинге Excel-файла: {e}")
        import traceback
        traceback.print_exc()
        return {}

def extract_time_from_cell(cell):
    """
    Извлекает время из ячейки Excel в формате 'HH:MM'.
    
    Args:
        cell: Ячейка Excel
        
    Returns:
        str: Время в формате 'HH:MM'
    """
    cell_value = cell.value
    
    # Если это datetime.time объект
    if hasattr(cell_value, 'hour') and hasattr(cell_value, 'minute'):
        return f"{cell_value.hour:02d}:{cell_value.minute:02d}"
    
    # Если это уже datetime, форматируем его
    if isinstance(cell_value, datetime):
        return cell_value.strftime('%H:%M')
    
    # Если это число, считаем его как время Excel (дробная часть дня)
    if isinstance(cell_value, (int, float)):
        # Конвертируем время Excel в часы и минуты
        total_minutes = int(cell_value * 24 * 60)
        hours = total_minutes // 60
        minutes = total_minutes % 60
        return f"{hours:02d}:{minutes:02d}"
    
    # Для строковых значений, которые выглядят как время (HH:MM:SS)
    if isinstance(cell_value, str) and ':' in cell_value:
        parts = cell_value.split(':')
        if len(parts) >= 2:
            try:
                hours = int(parts[0])
                minutes = int(parts[1])
                return f"{hours:02d}:{minutes:02d}"
            except ValueError:
                pass
    
    # Возвращаем значение как есть в случае неудачи
    return cell_value

def process_room_name(room, building):
    """
    Обрабатывает название кабинета и извлекает информацию о здании.
    
    В новом формате названия кабинетов имеют вид "VK.07" или "K2.3", где:
    - Первые символы могут указывать на здание (V для Villa, K для Kolibri)
    - После них следует название кабинета
    
    Args:
        room (str): Название кабинета
        building (str): Исходное название здания
        
    Returns:
        dict: Словарь с информацией о кабинете
    """
    if not room:
        return {
            "full_name": room,
            "display_name": room,
            "building": building
        }
    
    # Словарь сопоставлений префиксов кабинетов и зданий
    building_mappings = {
        "V": "Villa",
        "K": "Kolibri"
    }
    
    # Определяем здание по префиксу кабинета, если он присутствует
    detected_building = building
    display_name = room
    
    for prefix, building_name in building_mappings.items():
        if room.startswith(prefix):
            display_name = room[len(prefix):]  # Убираем префикс здания из названия кабинета
            detected_building = building_name
            break
    
    # Если здание не определено по префиксу, используем указанное здание
    if not detected_building and building:
        detected_building = building
    
    return {
        "full_name": room,
        "display_name": display_name,
        "building": detected_building
    }
