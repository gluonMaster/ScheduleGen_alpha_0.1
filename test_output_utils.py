from openpyxl import load_workbook

from output_utils import export_to_excel, make_safe_sheet_name


def test_make_safe_sheet_name_replaces_invalid_excel_chars_and_deduplicates():
    used = set()

    first = make_safe_sheet_name("G", "Musikunterricht / T. Liakhina", used)
    second = make_safe_sheet_name("G", "Musikunterricht ? T. Liakhina", used)

    assert first == "G_Musikunterricht _ T. Liakhina"
    assert second == "G_Musikunterricht _ T. Liakhi_2"
    assert len(first) <= 31
    assert len(second) <= 31


def test_export_to_excel_accepts_group_names_with_slash(tmp_path):
    class OptimizerStub:
        solution = [
            {
                "subject": "Vermietung",
                "group": "Musikunterricht / T. Liakhina",
                "teacher": "Olesya Fridel",
                "room": "1.09",
                "building": "Villa",
                "day": "Do",
                "start_time": "14:30",
                "end_time": "15:15",
                "duration": 45,
            }
        ]
        teachers = ["Olesya Fridel"]
        groups = ["Musikunterricht / T. Liakhina"]
        rooms = ["1.09"]

    output_file = tmp_path / "schedule.xlsx"

    assert export_to_excel(OptimizerStub(), filename=str(output_file)) is True

    workbook = load_workbook(output_file, read_only=True)
    assert "G_Musikunterricht _ T. Liakhina" in workbook.sheetnames
