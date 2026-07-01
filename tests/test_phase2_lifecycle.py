import hashlib
import importlib
import json
import sys
import zipfile

import pytest
from openpyxl import load_workbook

from gear_xls import backup_manager
from gear_xls import base_schedule_manager
from gear_xls import group_occupancy_snapshot
from gear_xls import integration
from gear_xls import lock_manager
from gear_xls import restore_manager
from gear_xls import schedule_mutation_coordinator
from gear_xls import state_manager
from gear_xls.excel_exporter import (
    ExcelExportValidationError,
    create_excel_from_html_data,
)
from gear_xls.schedule_state_errors import OccupancyUnavailable


MINIMAL_SCHEDULE_HTML = (
    "<html><head><style>.schedule-container{display:block}</style></head>"
    '<body><div id="menuDropdown"></div><div class="schedule-container"></div></body></html>'
)


def _write_json(path, payload):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _read_json(path):
    return json.loads(path.read_text(encoding="utf-8"))


def _project_root(tmp_path):
    root = tmp_path
    (root / "gear_xls" / "html_output").mkdir(parents=True)
    (root / "gear_xls" / "schedule_state").mkdir(parents=True)
    (root / "spiski").mkdir()
    (root / "xlsx_initial").mkdir()
    (root / "visualiser").mkdir()
    (root / "gui.py").write_text("", encoding="utf-8")
    (root / "gear_xls" / "server_routes.py").write_text("", encoding="utf-8")
    (root / "gear_xls" / "html_output" / "schedule.html").write_text(
        MINIMAL_SCHEDULE_HTML,
        encoding="utf-8",
    )
    for filename in backup_manager.ALLOWED_SPISKI_FILENAMES:
        (root / "spiski" / filename).write_text("item\n", encoding="utf-8")
    return root


def _event(block_id="event-1", **overrides):
    block = {
        "id": block_id,
        "day": "Mo",
        "building": "Villa",
        "room": "0.04",
        "start_time": "10:00",
        "end_time": "11:00",
        "lesson_type": "veranstaltung",
        "subject": "Veranstaltung",
        "teacher": "Event Manager",
        "students": "Parents",
        "color": "#7c3aed",
        "event_dates": ["2026-07-10"],
        "created_by": "em1",
        "created_by_name": "Event Manager",
        "owner_kind": "event_manager",
        "version": 3,
    }
    block.update(overrides)
    return block


def _group_block(**overrides):
    block = {
        "id": "group-1",
        "day": "Mo",
        "building": "Villa",
        "room": "0.04",
        "start_time": "10:30",
        "end_time": "11:30",
        "lesson_type": "group",
        "subject": "Math",
        "teacher": "Teacher",
        "students": "2A",
    }
    block.update(overrides)
    return block


def _snapshot(*blocks, generation_id="gen-1", source="test"):
    return {
        "schema_version": 1,
        "generation_id": generation_id,
        "generated_at": "2026-07-01T10:00:00Z",
        "source": source,
        "blocks": list(blocks),
    }


def _base_state(*blocks, published_at=None):
    return {"published_at": published_at, "published_by": "admin" if published_at else None, "blocks": list(blocks)}


def _individual_state(*blocks):
    return {"last_modified": "state-1", "blocks": list(blocks)}


def _login(client, role, login=None, display_name=None):
    with client.session_transaction() as session:
        session["login"] = login or role
        session["display_name"] = display_name or (login or role).title()
        session["role"] = role


def _patch_state_paths(monkeypatch, tmp_path):
    individual_path = tmp_path / "individual_lessons.json"
    base_path = tmp_path / "base_schedule.json"
    snapshot_path = tmp_path / "group_occupancy_snapshot.json"
    coordinator_path = tmp_path / "schedule_mutation.lock"
    monkeypatch.setattr(state_manager, "INDIVIDUAL_LESSONS_PATH", str(individual_path))
    monkeypatch.setattr(state_manager, "INDIVIDUAL_LOCK_PATH", str(individual_path) + ".lock")
    monkeypatch.setattr(state_manager, "SCHEDULE_HTML_PATH", str(tmp_path / "missing_schedule.html"))
    monkeypatch.setattr(base_schedule_manager, "BASE_SCHEDULE_PATH", str(base_path))
    monkeypatch.setattr(base_schedule_manager, "BASE_LOCK_PATH", str(base_path) + ".lock")
    top_level_base = sys.modules.get("base_schedule_manager")
    if top_level_base is not None:
        monkeypatch.setattr(top_level_base, "BASE_SCHEDULE_PATH", str(base_path))
        monkeypatch.setattr(top_level_base, "BASE_LOCK_PATH", str(base_path) + ".lock")
    monkeypatch.setattr(group_occupancy_snapshot, "GROUP_OCCUPANCY_SNAPSHOT_PATH", str(snapshot_path))
    monkeypatch.setattr(schedule_mutation_coordinator, "SCHEDULE_MUTATION_LOCK_PATH", str(coordinator_path))
    return individual_path, base_path, snapshot_path


def test_server_export_filters_events_and_rejects_event_only(tmp_path):
    output = tmp_path / "mixed.xlsx"
    regular = {
        "subject": "Deutsch",
        "students": "S",
        "teacher": "T",
        "room": "1.01",
        "building": "Villa",
        "day": "Mo",
        "start_time": "09:00",
        "end_time": "10:00",
        "duration": 60,
        "lesson_type": "individual",
    }
    event = {**regular, "subject": "Veranstaltung", "lesson_type": "veranstaltung"}
    subject_only_event = {**regular, "subject": "Veranstaltung", "lesson_type": "group"}

    with pytest.raises(ExcelExportValidationError) as excinfo:
        create_excel_from_html_data([event, subject_only_event], str(tmp_path / "events.xlsx"))

    assert excinfo.value.code == "NO_EXPORTABLE_ROWS"

    create_excel_from_html_data([event, regular, subject_only_event], str(output))
    ws = load_workbook(output).active

    assert ws.max_row == 2
    assert ws.cell(row=2, column=1).value == "Deutsch"
    assert ws.cell(row=2, column=10).value == "individual"


def test_client_export_contains_veranstaltung_skip():
    js = (backup_manager.os.path.dirname(backup_manager.__file__))
    path = backup_manager.os.path.join(js, "js_modules", "export_to_excel.js")
    text = open(path, "r", encoding="utf-8").read()

    assert "veranstaltung" in text
    assert "Veranstaltung" in text
    assert "scheduleData.push(activity)" in text


def test_publish_rejects_group_conflict_with_saved_event(tmp_path, monkeypatch):
    routes = importlib.import_module("gear_xls.server_routes")
    routes.app.config.update(TESTING=True)
    individual_path = tmp_path / "individual_lessons.json"
    base_path = tmp_path / "base_schedule.json"
    monkeypatch.setattr(routes.state_manager, "INDIVIDUAL_LESSONS_PATH", str(individual_path))
    monkeypatch.setattr(routes.state_manager, "INDIVIDUAL_LOCK_PATH", str(individual_path) + ".lock")
    monkeypatch.setattr(routes.state_manager, "SCHEDULE_HTML_PATH", str(tmp_path / "missing_schedule.html"))
    top_level_base = sys.modules.get("base_schedule_manager")
    if top_level_base is not None:
        monkeypatch.setattr(top_level_base, "BASE_SCHEDULE_PATH", str(base_path))
        monkeypatch.setattr(top_level_base, "BASE_LOCK_PATH", str(base_path) + ".lock")
    _write_json(individual_path, _individual_state(_event()))
    _write_json(base_path, _base_state(published_at=None))

    monkeypatch.setattr(routes.lock_manager, "get_lock_status", lambda: {"holder": "admin", "version": 1})
    monkeypatch.setattr(
        routes.restore_manager,
        "get_restore_status",
        lambda: {"active": False, "recovery_required": False, "generation": 0},
    )

    with routes.app.test_client() as client:
        _login(client, "admin", "admin", "Admin")
        response = client.post(
            "/api/schedule/publish",
            json={"blocks": [_group_block()], "expected_base_revision": None},
        )

    assert response.status_code == 409
    assert response.get_json()["code"] == "EVENT_ROOM_CONFLICT"
    assert _read_json(base_path)["blocks"] == []


def test_backup_v2_includes_snapshot_and_validates_events(tmp_path):
    root = _project_root(tmp_path)
    event = _event(event_dates=[], version=5)
    _write_json(root / "gear_xls" / "schedule_state" / "base_schedule.json", _base_state(published_at=None))
    _write_json(root / "gear_xls" / "schedule_state" / "individual_lessons.json", _individual_state(event))
    _write_json(root / "gear_xls" / "schedule_state" / "group_occupancy_snapshot.json", _snapshot())

    backup = backup_manager.create_backup("admin", "Admin", project_root=str(root))
    path = backup_manager.get_backup_path(backup["id"], str(root))

    with zipfile.ZipFile(path) as archive:
        manifest = json.loads(archive.read("manifest.json").decode("utf-8"))
        individual = json.loads(archive.read("state/individual_lessons.json").decode("utf-8"))
        snapshot_bytes = archive.read("state/group_occupancy_snapshot.json")

    assert manifest["schema_version"] == 2
    assert manifest["occupancy_snapshot"]["generation_id"] == "gen-1"
    assert manifest["occupancy_snapshot"]["sha256"] == hashlib.sha256(snapshot_bytes).hexdigest()
    assert individual["blocks"][0]["lesson_type"] == "veranstaltung"
    backup_manager.validate_individual_state(individual)

    invalid = _individual_state({**event, "version": 0})
    with pytest.raises(backup_manager.BackupValidationError):
        backup_manager.validate_individual_state(invalid)


def test_restore_v2_preserves_event_and_snapshot(tmp_path):
    root = _project_root(tmp_path)
    event = _event(version=7, event_dates=["2026-07-03"])
    _write_json(root / "gear_xls" / "schedule_state" / "base_schedule.json", _base_state(published_at=None))
    _write_json(root / "gear_xls" / "schedule_state" / "individual_lessons.json", _individual_state(event))
    _write_json(root / "gear_xls" / "schedule_state" / "group_occupancy_snapshot.json", _snapshot(generation_id="gen-restore"))

    backup = backup_manager.create_backup("admin", "Admin", project_root=str(root))
    _write_json(root / "gear_xls" / "schedule_state" / "individual_lessons.json", _individual_state())
    _write_json(root / "gear_xls" / "schedule_state" / "group_occupancy_snapshot.json", _snapshot(generation_id="stale"))

    lock_manager.acquire_lock("admin", project_root=str(root))
    result = restore_manager.restore_backup(backup["id"], "admin", "Admin", project_root=str(root))

    restored_individual = _read_json(root / "gear_xls" / "schedule_state" / "individual_lessons.json")
    restored_snapshot = _read_json(root / "gear_xls" / "schedule_state" / "group_occupancy_snapshot.json")
    assert result["ok"] is True
    assert restored_individual["blocks"][0]["version"] == 7
    assert restored_individual["blocks"][0]["event_dates"] == ["2026-07-03"]
    assert restored_individual["blocks"][0]["owner_kind"] == "event_manager"
    assert restored_snapshot["generation_id"] == "gen-restore"


def test_v1_restore_without_base_removes_stale_snapshot_and_blocks_events(tmp_path, monkeypatch):
    root = _project_root(tmp_path)
    backup_dir = root / "gear_xls" / "backups"
    backup_dir.mkdir()
    backup_id = "schedgen_backup_20260701_100000_deadbeef"
    entries = _v1_backup_entries(root)
    with zipfile.ZipFile(backup_dir / f"{backup_id}.zip", "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for name, data in entries.items():
            archive.writestr(name, data)
    _write_json(root / "gear_xls" / "schedule_state" / "group_occupancy_snapshot.json", _snapshot(generation_id="stale"))

    lock_manager.acquire_lock("admin", project_root=str(root))
    restore_manager.restore_backup(backup_id, "admin", "Admin", project_root=str(root))

    assert not (root / "gear_xls" / "schedule_state" / "group_occupancy_snapshot.json").exists()

    individual_path, base_path, snapshot_path = _patch_state_paths(monkeypatch, root / "gear_xls" / "schedule_state")
    monkeypatch.setattr(state_manager, "get_base_schedule", lambda: _base_state(published_at=None))
    assert individual_path.exists()
    assert base_path.exists()
    assert not snapshot_path.exists()
    with pytest.raises(OccupancyUnavailable):
        state_manager.create_event(
            {
                "day": "Mo",
                "building": "Villa",
                "room": "0.04",
                "start_time": "10:00",
                "end_time": "11:00",
                "students": "Parents",
                "event_dates": [],
            },
            {"login": "em1", "display_name": "Event Manager", "role": "event_manager"},
            {"login": "em1", "display_name": "Event Manager", "owner_kind": "event_manager"},
        )


def _v1_backup_entries(root):
    base = _base_state(published_at=None)
    individual = _individual_state()
    entries = {
        "state/base_schedule.json": json.dumps(base).encode("utf-8"),
        "state/individual_lessons.json": json.dumps(individual).encode("utf-8"),
        "html/schedule.html": MINIMAL_SCHEDULE_HTML.encode("utf-8"),
    }
    for archive_path in backup_manager.SPISKI_ARCHIVE_PATHS:
        entries[archive_path] = b"item\n"
    manifest = {
        "schema": backup_manager.BACKUP_SCHEMA,
        "schema_version": 1,
        "backup_kind": "manual",
        "created_at": "2026-07-01T10:00:00Z",
        "created_by": "admin",
        "created_by_display_name": "Admin",
        "comment": "",
        "project_root_id": backup_manager.get_project_root_id(str(root)),
        "app": "Kolibri SchedGen",
        "source": "web_editor_persisted_state",
        "dirty_dom_included": False,
        "base_revision": None,
        "individual_revision": "state-1",
        "includes": {
            "schedule_html": True,
            "base_schedule": True,
            "individual_lessons": True,
            "spiski": True,
            "lock_state": False,
            "restore_status": False,
            "source_excel": False,
        },
        "spiski_files": list(backup_manager.ALLOWED_SPISKI_FILENAMES),
        "files": [
            {"path": path, "sha256": hashlib.sha256(entries[path]).hexdigest(), "size": len(entries[path])}
            for path in backup_manager.V1_EXPECTED_CONTENT_PATHS
        ],
    }
    return {"manifest.json": json.dumps(manifest).encode("utf-8"), **entries}


def test_reset_preserves_events_replaces_old_individual_and_writes_snapshot(tmp_path, monkeypatch):
    state_dir = tmp_path / "state"
    html_dir = tmp_path / "html"
    html_dir.mkdir()
    staged_html = tmp_path / "staged_schedule.html"
    staged_html.write_text("<html>new schedule</html>", encoding="utf-8")
    individual_path, _base_path, snapshot_path = _patch_state_paths(monkeypatch, state_dir)
    monkeypatch.setattr(integration, "get_schedule_state_dir", lambda: str(state_dir))
    monkeypatch.setattr(integration, "get_html_output_dir", lambda: str(html_dir))
    old_individual = {
        "id": "old",
        "day": "Di",
        "building": "Villa",
        "room": "1.01",
        "start_time": "12:00",
        "end_time": "13:00",
        "lesson_type": "individual",
        "subject": "Old",
    }
    new_trial = {
        "id": "trial-new",
        "day": "Mo",
        "building": "Villa",
        "room": "1.02",
        "start_time": "09:00",
        "end_time": "10:00",
        "lesson_type": "trial",
        "subject": "Trial",
        "trial_dates": ["2026-07-06"],
    }
    _write_json(individual_path, _individual_state(_event(), old_individual))

    integration.reset_web_editor_state(
        [new_trial],
        group_buildings={"Villa": {"Mo": [_group_block(room="0.06", start_time="11:00", end_time="12:00")]}},
        snapshot_source="source.xlsx",
        html_source_path=str(staged_html),
    )

    state = _read_json(individual_path)
    assert [block["id"] for block in state["blocks"]] == ["trial-new", "event-1"]
    assert state["blocks"][1]["version"] == 3
    assert snapshot_path.exists()
    assert _read_json(snapshot_path)["source"] == "source.xlsx"
    assert (html_dir / "schedule.html").read_text(encoding="utf-8") == "<html>new schedule</html>"


def test_reset_aborts_when_generated_occupancy_conflicts_with_preserved_event(tmp_path, monkeypatch):
    state_dir = tmp_path / "state"
    html_dir = tmp_path / "html"
    html_dir.mkdir()
    (html_dir / "schedule.html").write_text("<html>old schedule</html>", encoding="utf-8")
    staged_html = tmp_path / "staged_conflict.html"
    staged_html.write_text("<html>new schedule</html>", encoding="utf-8")
    individual_path, _base_path, snapshot_path = _patch_state_paths(monkeypatch, state_dir)
    monkeypatch.setattr(integration, "get_schedule_state_dir", lambda: str(state_dir))
    monkeypatch.setattr(integration, "get_html_output_dir", lambda: str(html_dir))
    original = _individual_state(_event())
    _write_json(individual_path, original)
    _write_json(snapshot_path, _snapshot(generation_id="old"))

    with pytest.raises(integration.RegenerationEventConflict):
        integration.reset_web_editor_state(
            [],
            group_buildings={"Villa": {"Mo": [_group_block(start_time="10:30", end_time="11:30")]}},
            snapshot_source="source.xlsx",
            html_source_path=str(staged_html),
        )

    assert _read_json(individual_path) == original
    assert _read_json(snapshot_path)["generation_id"] == "old"
    assert (html_dir / "schedule.html").read_text(encoding="utf-8") == "<html>old schedule</html>"
