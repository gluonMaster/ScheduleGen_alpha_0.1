import json
import sys
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from gear_xls import backup_manager
from gear_xls import excel_exporter
from gear_xls import integration
from gear_xls import rooms_report
from gear_xls import rooms_routes
from gear_xls import schedule_structure
from gear_xls import state_manager
from gear_xls.services import schedule_pipeline


def _trial_block(day="So", trial_dates=None):
    return {
        "building": "Villa",
        "day": day,
        "room": "1.01",
        "subject": "Trial",
        "teacher": "Teacher",
        "students": "Student",
        "start_time": "10:00",
        "end_time": "11:00",
        "lesson_type": "trial",
        "trial_dates": trial_dates if trial_dates is not None else ["2026-05-24"],
    }


def _regular_block(day="So", lesson_type="individual"):
    return {
        "building": "Villa",
        "day": day,
        "room": "1.01",
        "subject": "Deutsch",
        "teacher": "Teacher",
        "students": "Student",
        "start_time": "10:00",
        "end_time": "11:00",
        "lesson_type": lesson_type,
    }


@pytest.fixture()
def isolated_individual_state(tmp_path, monkeypatch):
    path = tmp_path / "individual_lessons.json"
    monkeypatch.setattr(state_manager, "INDIVIDUAL_LESSONS_PATH", str(path))
    monkeypatch.setattr(state_manager, "INDIVIDUAL_LOCK_PATH", str(path) + ".lock")
    return path


def test_state_manager_accepts_sunday_trial_with_sunday_date():
    block = _trial_block()

    assert state_manager._validate_block(block, "admin") is None
    assert block["trial_dates"] == ["2026-05-24"]


def test_state_manager_rejects_sunday_trial_with_non_sunday_date():
    block = _trial_block(trial_dates=["2026-05-25"])

    assert "not matching block day So" in state_manager._validate_block(block, "admin")


@pytest.mark.parametrize("role", ["admin", "editor", "organizer"])
def test_state_manager_rejects_non_trial_sunday_for_all_roles(role):
    block = _regular_block(lesson_type="individual" if role != "organizer" else "trial")
    if role == "organizer":
        block["lesson_type"] = "individual"

    assert state_manager._validate_block(block, role)


def test_convert_sunday_trial_to_regular_is_rejected(isolated_individual_state):
    created, error = state_manager.add_block(_trial_block(), "admin")

    assert error is None
    converted, convert_error = state_manager.convert_block_to_regular(created["id"], "admin")
    assert converted is None
    assert convert_error == "Sunday is allowed only for trial lessons"


def test_backup_validation_accepts_sunday_trial_individual_state():
    backup_manager.validate_individual_state(
        {"last_modified": None, "blocks": [{"id": "trial-1", **_trial_block()}]}
    )


def test_backup_validation_rejects_sunday_trial_with_wrong_weekday():
    with pytest.raises(backup_manager.BackupValidationError):
        backup_manager.validate_individual_state(
            {
                "last_modified": None,
                "blocks": [{"id": "trial-1", **_trial_block(trial_dates=["2026-05-25"])}],
            }
        )


def test_backup_validation_rejects_non_trial_sunday_individual_state():
    with pytest.raises(backup_manager.BackupValidationError):
        backup_manager.validate_individual_state(
            {"last_modified": None, "blocks": [{"id": "lesson-1", **_regular_block()}]}
        )


def test_backup_validation_rejects_sunday_base_state():
    with pytest.raises(backup_manager.BackupValidationError):
        backup_manager.validate_base_state(
            {"published_at": None, "published_by": None, "blocks": [_regular_block(lesson_type="group")]}
        )


def test_excel_export_keeps_sunday_trial_row(tmp_path):
    output = tmp_path / "sunday_trial.xlsx"
    row = {
        "subject": "Trial",
        "students": "Student",
        "teacher": "Teacher",
        "room": "1.01",
        "building": "Villa",
        "day": "So",
        "start_time": "10:00",
        "end_time": "11:00",
        "duration": 60,
        "lesson_type": "trial",
        "trial_dates_json": json.dumps(["2026-05-24"]),
    }

    excel_exporter.create_excel_from_html_data([row], str(output))
    ws = load_workbook(output).active

    assert ws.cell(row=2, column=6).value == "So"
    assert ws.cell(row=2, column=10).value == "trial"


def test_excel_export_validation_rejects_non_trial_sunday():
    with pytest.raises(excel_exporter.ExcelExportValidationError) as excinfo:
        excel_exporter.validate_schedule_data_for_export([_regular_block(lesson_type="group")])

    assert excinfo.value.code == "SUNDAY_REGULAR_FORBIDDEN"


def test_schedule_structure_binds_empty_sunday_column_to_building_room():
    buildings = schedule_structure.build_schedule_structure(
        {
            "a1": {
                "building": "Villa",
                "day": "Mo",
                "room": "1.02",
                "room_display": "1.02",
                "subject": "Deutsch",
                "teacher": "Teacher",
                "students": "Group",
                "start_time": "10:00",
                "end_time": "11:00",
            }
        }
    )

    villa = buildings["Villa"]
    assert villa["_default_room"] == "1.02"
    assert villa["_rooms"]["So"] == ["1.02"]
    assert villa["_max_cols"]["So"] == 1


def test_pipeline_strips_non_group_blocks_from_generated_html():
    html = """
    <div class='activity-block lesson-type-group' data-lesson-type='group'>G</div>
    <div class='activity-block lesson-type-trial' data-lesson-type='trial'>T</div>
    <div class='activity-block lesson-type-nachhilfe' data-lesson-type='nachhilfe'>N</div>
    """

    stripped, removed = schedule_pipeline.strip_non_group_activity_blocks_from_html(html)

    assert removed == 2
    assert "data-lesson-type='group'" in stripped
    assert "data-lesson-type='trial'" not in stripped
    assert "lesson-type-nachhilfe" not in stripped


def test_pipeline_collects_trial_blocks_for_individual_state():
    buildings = schedule_structure.build_schedule_structure(
        {
            "group-1": {
                "building": "Villa",
                "day": "Mo",
                "room": "1.01",
                "room_display": "1.01",
                "subject": "Deutsch",
                "teacher": "Teacher",
                "students": "Group",
                "start_time": "10:00",
                "end_time": "11:00",
            },
            "trial-1": {
                "building": "Kolibri",
                "day": "So",
                "room": "2.1",
                "room_display": "2.1",
                "subject": "Trial",
                "teacher": "Teacher",
                "students": "Student",
                "start_time": "09:00",
                "end_time": "10:00",
                "lesson_type": "trial",
                "trial_dates": ["2026-05-24"],
            },
        }
    )

    blocks = schedule_pipeline.collect_individual_blocks_from_buildings(buildings)

    assert len(blocks) == 1
    assert blocks[0]["id"] == "trial-1"
    assert blocks[0]["lesson_type"] == "trial"
    assert blocks[0]["day"] == "So"
    assert blocks[0]["building"] == "Kolibri"
    assert blocks[0]["room"] == "2.1"
    assert blocks[0]["start_time"] == "09:00"
    assert blocks[0]["trial_dates"] == ["2026-05-24"]


def test_reset_web_editor_state_initializes_individual_blocks(tmp_path, monkeypatch):
    block = {"id": "trial-1", **_trial_block()}

    monkeypatch.setattr(integration, "get_schedule_state_dir", lambda: str(tmp_path))

    integration.reset_web_editor_state([block])

    base_state = json.loads((tmp_path / "base_schedule.json").read_text(encoding="utf-8"))
    individual_state = json.loads((tmp_path / "individual_lessons.json").read_text(encoding="utf-8"))

    assert base_state["blocks"] == []
    assert individual_state["blocks"] == [block]
    assert individual_state["last_modified"]


def test_rooms_report_includes_sunday_trial_blocks(tmp_path, monkeypatch):
    base_path = tmp_path / "base_schedule.json"
    individual_path = tmp_path / "individual_lessons.json"
    base_path.write_text(json.dumps({"published_at": "2026-05-21T10:00:00", "blocks": []}), encoding="utf-8")
    individual_path.write_text(json.dumps({"blocks": [_trial_block()]}), encoding="utf-8")

    monkeypatch.setattr(rooms_report, "BASE_SCHEDULE_PATH", str(base_path))
    monkeypatch.setattr(rooms_report, "INDIVIDUAL_LESSONS_PATH", str(individual_path))
    monkeypatch.setattr(rooms_report, "_load_configured_rooms", lambda: {"Villa": ["1.01"]})

    data = rooms_report.compute_availability()

    sunday_slots = data["buildings"]["Villa"]["days"]["So"]["1.01"]
    assert rooms_report.DAY_ORDER["So"] == 6
    assert sunday_slots[0]["lesson_type"] == "trial"


def test_rooms_page_exposes_sunday_day_filter():
    js = (PROJECT_ROOT / "gear_xls" / "static" / "rooms_report.js").read_text(encoding="utf-8")

    assert 'id="day-filter-group"' in rooms_routes.ROOMS_PAGE_TEMPLATE
    assert 'id="d-So"' in rooms_routes.ROOMS_PAGE_TEMPLATE
    assert 'label for="d-So">So' in rooms_routes.ROOMS_PAGE_TEMPLATE
    assert '["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]' in js
    assert "ensureDayFilterControls" in js


def test_rooms_page_hides_availability_table_by_default():
    template = rooms_routes.ROOMS_PAGE_TEMPLATE

    assert 'id="btn-toggle-rooms-table"' in template
    assert "Показать таблицу занятости аудиторий" in template
    assert 'id="rooms-navigation-status" hidden' in template
    assert '<div id="rooms-table-wrap" hidden>' in template


def test_rooms_available_results_link_to_schedule_preflight():
    js = (PROJECT_ROOT / "gear_xls" / "static" / "rooms_report.js").read_text(encoding="utf-8")

    assert "available-room-link" in js
    assert "data-building" in js
    assert "data-room" in js
    assert "data-day" in js
    assert "data-start" in js
    assert "data-end" in js
    assert 'params.set("rooms_nav", "1")' in js
    assert 'lockApiRequest("/api/lock/acquire", "POST")' in js
    assert 'lockApiRequest("/api/lock", "DELETE")' in js


def test_schedule_loads_rooms_focus_script():
    server_routes = (PROJECT_ROOT / "gear_xls" / "server_routes.py").read_text(encoding="utf-8")
    focus_js = (PROJECT_ROOT / "gear_xls" / "static" / "rooms_schedule_focus.js").read_text(encoding="utf-8")
    search_js = (PROJECT_ROOT / "gear_xls" / "static" / "schedule_search_ui.js").read_text(encoding="utf-8")
    generator = (PROJECT_ROOT / "gear_xls" / "html_javascript.py").read_text(encoding="utf-8")

    assert "/static/rooms_schedule_focus.js" in server_routes
    assert "rooms_schedule_focus.js?v=20260521_3" in search_js
    assert "/api/columns" in focus_js
    assert "addColumnIfMissing" in focus_js
    assert "toggleDay" in focus_js
    assert "rooms-schedule-focus-cell" in focus_js
    assert "window.gridStart = gridStart" in generator
    assert "window.timeInterval = timeInterval" in generator


def test_rooms_focus_clears_temporary_highlight_on_editor_actions():
    focus_js = (PROJECT_ROOT / "gear_xls" / "static" / "rooms_schedule_focus.js").read_text(encoding="utf-8")
    auth_js = (PROJECT_ROOT / "gear_xls" / "static" / "auth_ui.js").read_text(encoding="utf-8")

    assert "MutationObserver" in focus_js
    assert "blockMatchesActiveFocus" in focus_js
    assert 'target.closest(".toggle-day-button")' in focus_js
    assert 'event.key === "Escape"' in focus_js
    assert '"schedgen:edit-mode-change"' in focus_js
    assert "RoomsScheduleFocus" in focus_js
    assert 'new CustomEvent("schedgen:edit-mode-change"' in auth_js


def test_final_visualiser_filters_sunday_rows():
    from visualiser import data_processor as visualiser_data

    df = pd.DataFrame(
        [
            {"day": "Mo", "subject": "Math", "group": "G1", "teacher": "T", "room": "1.01", "building": "Villa", "start_time": "10:00", "end_time": "11:00", "duration": 60},
            {"day": "So", "subject": "Trial", "group": "S1", "teacher": "T", "room": "1.02", "building": "Villa", "start_time": "10:00", "end_time": "11:00", "duration": 60},
        ]
    )

    days, schedule_by_day = visualiser_data.process_schedule_data(df)

    assert days == ["Mo"]
    assert list(schedule_by_day) == ["Mo"]


def test_final_tv_visualiser_filters_sunday_rows():
    from visualiserTV import data_processor as visualiser_tv_data

    df = pd.DataFrame(
        [
            {"day": "Sa", "subject": "Math", "group": "G1", "teacher": "T", "room": "1.01", "building": "Villa", "start_time": "10:00", "end_time": "11:00", "duration": 60},
            {"day": "So", "subject": "Trial", "group": "S1", "teacher": "T", "room": "1.02", "building": "Villa", "start_time": "10:00", "end_time": "11:00", "duration": 60},
        ]
    )

    days, schedule_by_day = visualiser_tv_data.process_schedule_data(df)

    assert days == ["Sa"]
    assert list(schedule_by_day) == ["Sa"]
